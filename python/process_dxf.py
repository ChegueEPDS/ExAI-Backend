'''import sys
import ezdxf
import re
import json
import pandas as pd
from openpyxl import Workbook
from math import hypot
import logging
import os

# 📒 Log konfigurálása
log_filename = "results/process.log"
os.makedirs("results", exist_ok=True)

logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# 📁 Fájl elérési út ellenőrzése
if len(sys.argv) < 2:
    logging.error("Hiba: nincs fájlmegadás.")
    print("Hiba: nincs fájlmegadás.", file=sys.stderr)
    sys.exit(1)

dxf_path = sys.argv[1]
logging.info(f"DXF fájl beolvasása: {dxf_path}")

try:
    doc = ezdxf.readfile(dxf_path)
except Exception as e:
    logging.exception(f"Hiba a DXF beolvasásakor: {e}")
    print(f"Hiba a DXF beolvasásakor: {e}", file=sys.stderr)
    sys.exit(1)

msp = doc.modelspace()
lines = []
texts = []
fittings = {}

# ➤ Közelség alapján DN hozzárendelése
def midpoint(p1, p2):
    return ((p1[0]+p2[0])/2, (p1[1]+p2[1])/2)

def to_xy_tuple(vec):
    return (float(vec[0]), float(vec[1]))

def distance_point_to_segment(px, py, x1, y1, x2, y2):
    """Visszaadja a (px, py) pont távolságát az (x1,y1)-(x2,y2) szakasztól"""
    dx = x2 - x1
    dy = y2 - y1
    if dx == dy == 0:  # a szakasz egy pont
        return hypot(px - x1, py - y1)

    t = ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)
    t = max(0, min(1, t))  # clamp 0–1
    nearest_x = x1 + t * dx
    nearest_y = y1 + t * dy
    return hypot(px - nearest_x, py - nearest_y)

MAX_DISTANCE = 500  # új távolsági küszöb

dn_pattern = re.compile(r"(?:DN|Ø|D\.?)\s*(\d+)", re.IGNORECASE)

# ➤ Vonalak beolvasása
for e in msp:
    if e.dxftype() == 'LINE':
        lines.append({
            "start": e.dxf.start,
            "end": e.dxf.end
        })
logging.info(f"{len(lines)} LINE objektum beolvasva.")

# ➤ Szövegek
for e in msp.query('TEXT MTEXT'):
    text_content = e.plain_text() if e.dxftype() == 'MTEXT' else e.dxf.text
    pos = to_xy_tuple(e.dxf.insert)
    
    # Keress DN-t, de akkor is mentsd a teljes szöveget
    match = dn_pattern.search(text_content)
    dn_text = f"DN{match.group(1)}" if match else None

    texts.append({
        "text": text_content,
        "position": pos,
        "dn": dn_text
    })
logging.info(f"{len(texts)} szövegobjektum beolvasva, köztük {sum(1 for t in texts if t['dn'])} DN-azonosított.")



pipe_data = []
matched_count = 0
unmatched_count = 0

for idx, line in enumerate(lines):
    positions = [
        midpoint(line['start'], line['end']),
        line['start'],
        line['end'],
    ]
    closest_dn = None
    min_dist = float('inf')

    for pos in positions:
        for t in texts:
            dist = distance_point_to_segment(
                t['position'][0], t['position'][1],
                line['start'][0], line['start'][1],
                line['end'][0], line['end'][1]
            )
            if dist < MAX_DISTANCE and dist < min_dist:
                min_dist = dist
                closest_dn = t['text']

    length = hypot(
        line['end'][0] - line['start'][0],
        line['end'][1] - line['start'][1]
    )

    if closest_dn:
        matched_count += 1
    else:
        unmatched_count += 1

    pipe_data.append({
        "id": idx,  # 🟢 most már működik
        "length": round(length, 2),
        "dn": closest_dn or "unknown",
        "start": to_xy_tuple(line['start']),
        "end": to_xy_tuple(line['end']),
        "midpoint": to_xy_tuple(midpoint(line['start'], line['end'])),
        "nearestText": closest_dn,
        "distanceToNearest": round(min_dist, 2) if closest_dn else None
    })

logging.info(f"{matched_count} csőszakaszhoz sikerült DN-t rendelni.")
logging.info(f"{unmatched_count} csőszakaszhoz nem talált DN szöveget (>= {MAX_DISTANCE} egység távolság).")
logging.info(f"{len(pipe_data)} csőszakasz feldolgozva.")

# 🔽 EZT MINDIG A for-CIKLUS UTÁN TEDD:
unknowns = [p for p in pipe_data if p['dn'] == 'unknown']
debug_path = f"results/debug_unknowns_{int(pd.Timestamp.now().timestamp())}.json"
with open(debug_path, 'w') as f:
    json.dump(unknowns, f, indent=2)

logging.info(f"{len(unknowns)} 'unknown' DN mentve ide: {debug_path}")

# ➤ Blokkok (szerelvények)
for e in msp.query('INSERT'):
    name = e.dxf.name.upper()
    for key in ["ELBOW", "T_IDOM", "VALVE"]:
        if key in name:
            fittings[key] = fittings.get(key, 0) + 1
logging.info(f"{len(fittings)} szerelvénytípus azonosítva: {fittings}")

# ➤ Excel export
wb = Workbook()
ws = wb.active
ws.title = "Pipes"
ws.append(["DN", "Length", "Start X", "Start Y", "End X", "End Y", "Midpoint X", "Midpoint Y"])
for p in pipe_data:
    ws.append([
        p["dn"],
        p["length"],
        round(p["start"][0], 2), round(p["start"][1], 2),
        round(p["end"][0], 2), round(p["end"][1], 2),
        round(p["midpoint"][0], 2), round(p["midpoint"][1], 2)
    ])

ws2 = wb.create_sheet("Fittings")
ws2.append(["Type", "Count"])
for k, v in fittings.items():
    ws2.append([k, v])

excel_path = f"results/output_{int(pd.Timestamp.now().timestamp())}.xlsx"
wb.save(excel_path)
logging.info(f"Excel mentve: {excel_path}")

def generate_svg(lines):
    from io import StringIO

    # 🔍 Legkisebb és legnagyobb x, y koordináta meghatározása
    all_coords = [coord for line in lines for coord in [line['start'], line['end']]]
    min_x = min(p[0] for p in all_coords)
    max_x = max(p[0] for p in all_coords)
    min_y = min(p[1] for p in all_coords)
    max_y = max(p[1] for p in all_coords)

    width = max_x - min_x
    height = max_y - min_y

    svg = StringIO()
    svg.write(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'viewBox="{min_x} {min_y} {width} {height}" style="background:#fff;">\n'
        f'<style>\n'
        f'.highlight {{ stroke: red; stroke-width: 12; }}\n'
        f'</style>\n'
    )

    for pipe in lines:
        x1, y1 = pipe['start']
        x2, y2 = pipe['end']
        pipe_id = pipe['id']
        svg.write(
            f'<line id="pipe-{pipe_id}" x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
            f'stroke="black" stroke-width="5" />\n'
        )

    svg.write('</svg>')
    return svg.getvalue()

# ➤ JSON kimenet
svg_content = generate_svg(pipe_data)

output = {
    "pipes": pipe_data,
    "fittings": [{"type": k, "count": v} for k, v in fittings.items()],
    "excel_path": excel_path,
    "log": log_filename,
    "dxfSvg": svg_content
}

print(json.dumps(output))

'''
import sys
import ezdxf
import re
import json
import pandas as pd
from openpyxl import Workbook
from math import hypot
import logging
import os

# 📒 Log konfigurálása
log_filename = "results/process.log"
os.makedirs("results", exist_ok=True)

logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# 📁 Fájl elérési út ellenőrzése
if len(sys.argv) < 2:
    logging.error("Hiba: nincs fájlmegadás.")
    print("Hiba: nincs fájlmegadás.", file=sys.stderr)
    sys.exit(1)

dxf_path = sys.argv[1]
logging.info(f"DXF fájl beolvasása: {dxf_path}")

try:
    doc = ezdxf.readfile(dxf_path)
except Exception as e:
    logging.exception(f"Hiba a DXF beolvasásakor: {e}")
    print(f"Hiba a DXF beolvasásakor: {e}", file=sys.stderr)
    sys.exit(1)

msp = doc.modelspace()
lines = []
texts = []
fittings = {}

# ➤ Közelség alapján DN hozzárendelése
def midpoint(p1, p2):
    return ((p1[0]+p2[0])/2, (p1[1]+p2[1])/2)

def to_xy_tuple(vec):
    return (float(vec[0]), float(vec[1]))

def distance_point_to_segment(px, py, x1, y1, x2, y2):
    """Visszaadja a (px, py) pont távolságát az (x1,y1)-(x2,y2) szakasztól"""
    dx = x2 - x1
    dy = y2 - y1
    if dx == dy == 0:  # a szakasz egy pont
        return hypot(px - x1, py - y1)

    t = ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)
    t = max(0, min(1, t))  # clamp 0–1
    nearest_x = x1 + t * dx
    nearest_y = y1 + t * dy
    return hypot(px - nearest_x, py - nearest_y)

MAX_DISTANCE = 500  # új távolsági küszöb

dn_pattern = re.compile(r"(?:DN|Ø|D\.?)\s*(\d+)", re.IGNORECASE)

# ➤ Vonalak beolvasása
for e in msp:
    if e.dxftype() == 'LINE':
        lines.append({
            "start": e.dxf.start,
            "end": e.dxf.end
        })
logging.info(f"{len(lines)} LINE objektum beolvasva.")

# ➤ Szövegek
for e in msp.query('TEXT MTEXT'):
    text_content = e.plain_text() if e.dxftype() == 'MTEXT' else e.dxf.text
    pos = to_xy_tuple(e.dxf.insert)
    
    # Keress DN-t, de akkor is mentsd a teljes szöveget
    match = dn_pattern.search(text_content)
    dn_text = f"DN{match.group(1)}" if match else None

    texts.append({
        "text": text_content,
        "position": pos,
        "dn": dn_text
    })
logging.info(f"{len(texts)} szövegobjektum beolvasva, köztük {sum(1 for t in texts if t['dn'])} DN-azonosított.")

pipe_data = []
matched_count = 0
unmatched_count = 0

for idx, line in enumerate(lines):
    positions = [
        midpoint(line['start'], line['end']),
        line['start'],
        line['end'],
    ]
    closest_dn = None
    min_dist = float('inf')

    for pos in positions:
        for t in texts:
            dist = distance_point_to_segment(
                t['position'][0], t['position'][1],
                line['start'][0], line['start'][1],
                line['end'][0], line['end'][1]
            )
            if dist < MAX_DISTANCE and dist < min_dist:
                min_dist = dist
                closest_dn = t['text']

    length = hypot(
        line['end'][0] - line['start'][0],
        line['end'][1] - line['start'][1]
    )

    if closest_dn:
        matched_count += 1
    else:
        unmatched_count += 1

    pipe_data.append({
        "id": idx,  # 🟢 most már működik
        "length": round(length, 2),
        "dn": closest_dn or "unknown",
        "start": to_xy_tuple(line['start']),
        "end": to_xy_tuple(line['end']),
        "midpoint": to_xy_tuple(midpoint(line['start'], line['end'])),
        "nearestText": closest_dn,
        "distanceToNearest": round(min_dist, 2) if closest_dn else None
    })

logging.info(f"{matched_count} csőszakaszhoz sikerült DN-t rendelni.")
logging.info(f"{unmatched_count} csőszakaszhoz nem talált DN szöveget (>= {MAX_DISTANCE} egység távolság).")
logging.info(f"{len(pipe_data)} csőszakasz feldolgozva.")

# 🔽 EZT MINDIG A for-CIKLUS UTÁN TEDD:
unknowns = [p for p in pipe_data if p['dn'] == 'unknown']
debug_path = f"results/debug_unknowns_{int(pd.Timestamp.now().timestamp())}.json"
with open(debug_path, 'w') as f:
    json.dump(unknowns, f, indent=2)

logging.info(f"{len(unknowns)} 'unknown' DN mentve ide: {debug_path}")

# ➤ Blokkok (szerelvények)
fittings_positions = []

for e in msp.query('INSERT'):
    name = e.dxf.name.upper()
    for key in ["ELBOW", "T_IDOM", "VALVE"]:
        if key in name:
            pos = to_xy_tuple(e.dxf.insert)
            fittings_positions.append({
                "type": key,
                "position": pos
            })
            fittings[key] = fittings.get(key, 0) + 1

logging.info(f"{len(fittings)} szerelvénytípus azonosítva: {fittings}")

# ➤ Excel export
wb = Workbook()
ws = wb.active
ws.title = "Pipes"
ws.append(["DN", "Length", "Start X", "Start Y", "End X", "End Y", "Midpoint X", "Midpoint Y"])
for p in pipe_data:
    ws.append([
        p["dn"],
        p["length"],
        round(p["start"][0], 2), round(p["start"][1], 2),
        round(p["end"][0], 2), round(p["end"][1], 2),
        round(p["midpoint"][0], 2), round(p["midpoint"][1], 2)
    ])

ws2 = wb.create_sheet("Fittings")
ws2.append(["Type", "Count"])
for k, v in fittings.items():
    ws2.append([k, v])

excel_path = f"results/output_{int(pd.Timestamp.now().timestamp())}.xlsx"
wb.save(excel_path)
logging.info(f"Excel mentve: {excel_path}")

def generate_svg(pipes, texts, fittings_positions):
    from io import StringIO
    from collections import defaultdict

    def compute_focus_bbox(positions, cell_size=10000, margin_cells=1):
        """Legsűrűbb cella körüli fókuszterület, extra margóval"""
        from collections import defaultdict

        grid = defaultdict(int)
        for (x, y) in positions:
            cx = int(x // cell_size)
            cy = int(y // cell_size)
            grid[(cx, cy)] += 1

        # 📍 Legzsúfoltabb cella
        focus_cx, focus_cy = max(grid.items(), key=lambda x: x[1])[0]

        min_x = (focus_cx - margin_cells) * cell_size
        max_x = (focus_cx + 1 + margin_cells) * cell_size
        min_y = (focus_cy - margin_cells) * cell_size
        max_y = (focus_cy + 1 + margin_cells) * cell_size

        return min_x, min_y, max_x, max_y

    # 🔍 Csak a valóban megjelenítendő elemek pozíciói
    positions = [p['start'] for p in pipes] + [p['end'] for p in pipes]
    positions += [t['position'] for t in texts]
    positions += [f['position'] for f in fittings_positions]

    # 📐 Fókuszbox számítás
    min_x, min_y, max_x, max_y = compute_focus_bbox(positions)
    width = max_x - min_x
    height = max_y - min_y

    # 🖼 SVG összerakása
    svg = StringIO()
    svg.write(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'viewBox="{min_x} {min_y} {width} {height}" style="background:#fff;">\n'
        f'<style>\n'
        f'.highlight {{ stroke: red; stroke-width: 12; }}\n'
        f'text {{ font-size: 200px; fill: blue; text-anchor: middle; }}\n'
        f'</style>\n'
    )

    # 🟢 Csövek
    svg.write(f'<g id="pipes">\n')
    for pipe in pipes:
        x1, y1 = pipe['start']
        x2, y2 = pipe['end']
        pipe_id = pipe['id']
        svg.write(
            f'<line id="pipe-{pipe_id}" x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
            f'stroke="black" stroke-width="5" />\n'
        )
    svg.write(f'</g>\n')

    # 🔵 Szövegek (csak a táblázatba kerültek – azaz DN vagy azonosított szövegek)
    svg.write(f'<g id="texts">\n')
    for t in texts:
        x, y = t['position']
        content = t['text'].replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        svg.write(
            f'<text x="{x}" y="{y}" dy=".3em">{content}</text>\n'
        )
    svg.write(f'</g>\n')

    # 🟢 Szerelvények (csak az azonosított típusok)
    svg.write(f'<g id="fittings">\n')

    label_map = {
        "ELBOW": "E",
        "T_IDOM": "T",
        "VALVE": "V"
    }

    for fitting in fittings_positions:
        x, y = fitting['position']
        label = label_map.get(fitting['type'], fitting['type'][0])  # <-- csak E, T, V
        svg.write(
            f'<circle cx="{x}" cy="{y}" r="200" fill="green" />\n'
            f'<text x="{x}" y="{y}" dy=".3em" fill="white">{label}</text>\n'
        )
    svg.write(f'</g>\n')

    svg.write('</svg>')
    return svg.getvalue()

# ➤ JSON kimenet
svg_content = generate_svg(pipe_data, [], fittings_positions) #texts

output = {
    "pipes": pipe_data,
    "fittings": [{"type": k, "count": v} for k, v in fittings.items()],
    "fittingPositions": fittings_positions,
   # "texts": texts,
    "excel_path": excel_path,
    "log": log_filename,
    "dxfSvg": svg_content
}

print(json.dumps(output))